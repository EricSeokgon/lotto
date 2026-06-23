"""SPEC-LOTTO-116: 당첨 번호 데이터 내보내기 테스트.

GET /api/export/csv  → CSV 파일 다운로드 (한국어 헤더, utf-8-sig)
GET /api/export/xlsx → Excel 파일 다운로드 (openpyxl)
"""
from __future__ import annotations

import datetime
import io
from unittest.mock import patch

import openpyxl
from fastapi.testclient import TestClient

from lotto.models import DrawResult
from lotto.web.app import app

client = TestClient(app)


# ---------------------------------------------------------------------------
# 테스트 픽스처 헬퍼
# ---------------------------------------------------------------------------

def _make_draw(drw_no: int, nums: list, bonus: int) -> DrawResult:
    """테스트용 DrawResult 생성."""
    return DrawResult(
        drwNo=drw_no,
        date=datetime.date(2002, 12, 7),
        n1=nums[0],
        n2=nums[1],
        n3=nums[2],
        n4=nums[3],
        n5=nums[4],
        n6=nums[5],
        bonus=bonus,
        prize1Amount=1000000000,
        prize1Winners=3,
    )


SAMPLE_DRAWS = [
    _make_draw(1, [10, 23, 29, 33, 37, 40], 16),
    _make_draw(2, [1, 2, 3, 4, 5, 6], 7),
]

EMPTY_DRAWS: list = []


# ---------------------------------------------------------------------------
# CSV 엔드포인트 테스트
# ---------------------------------------------------------------------------

def test_export_csv_returns_200() -> None:
    """GET /api/export/csv → 200 OK."""
    with patch("lotto.web.data.get_draws", return_value=SAMPLE_DRAWS):
        resp = client.get("/api/export/csv")
    assert resp.status_code == 200


def test_export_csv_content_type() -> None:
    """Content-Type이 text/csv를 포함한다."""
    with patch("lotto.web.data.get_draws", return_value=SAMPLE_DRAWS):
        resp = client.get("/api/export/csv")
    assert "text/csv" in resp.headers["content-type"]


def test_export_csv_content_disposition() -> None:
    """Content-Disposition 헤더에 lotto_draws.csv와 attachment가 포함된다."""
    with patch("lotto.web.data.get_draws", return_value=SAMPLE_DRAWS):
        resp = client.get("/api/export/csv")
    cd = resp.headers["content-disposition"]
    assert "lotto_draws.csv" in cd
    assert "attachment" in cd


def test_export_csv_has_header_row() -> None:
    """첫 번째 행에 한국어 열 헤더가 포함된다."""
    with patch("lotto.web.data.get_draws", return_value=SAMPLE_DRAWS):
        resp = client.get("/api/export/csv")
    # utf-8-sig BOM 처리
    text = resp.content.decode("utf-8-sig")
    first_line = text.splitlines()[0]
    assert "회차" in first_line
    assert "추첨일" in first_line
    assert "보너스" in first_line


def test_export_csv_row_count() -> None:
    """행 수 = 헤더 1행 + 회차 수."""
    with patch("lotto.web.data.get_draws", return_value=SAMPLE_DRAWS):
        resp = client.get("/api/export/csv")
    text = resp.content.decode("utf-8-sig")
    lines = [ln for ln in text.splitlines() if ln.strip()]
    assert len(lines) == len(SAMPLE_DRAWS) + 1


def test_export_csv_empty_draws() -> None:
    """회차가 없으면 헤더 행만 반환된다."""
    with patch("lotto.web.data.get_draws", return_value=EMPTY_DRAWS):
        resp = client.get("/api/export/csv")
    assert resp.status_code == 200
    text = resp.content.decode("utf-8-sig")
    lines = [ln for ln in text.splitlines() if ln.strip()]
    assert len(lines) == 1  # 헤더만


def test_export_csv_data_values() -> None:
    """데이터 행의 값이 올바르게 채워진다 (회차 번호, 날짜)."""
    with patch("lotto.web.data.get_draws", return_value=SAMPLE_DRAWS):
        resp = client.get("/api/export/csv")
    text = resp.content.decode("utf-8-sig")
    lines = text.splitlines()
    # drwNo=1 행 (정렬 후 첫 번째 데이터 행)
    row1 = lines[1]
    assert row1.startswith("1,")
    assert "2002-12-07" in row1


# ---------------------------------------------------------------------------
# Excel 엔드포인트 테스트
# ---------------------------------------------------------------------------

def test_export_xlsx_returns_200() -> None:
    """GET /api/export/xlsx → 200 OK."""
    with patch("lotto.web.data.get_draws", return_value=SAMPLE_DRAWS):
        resp = client.get("/api/export/xlsx")
    assert resp.status_code == 200


def test_export_xlsx_content_type() -> None:
    """Content-Type이 xlsx MIME 타입이다."""
    with patch("lotto.web.data.get_draws", return_value=SAMPLE_DRAWS):
        resp = client.get("/api/export/xlsx")
    assert "spreadsheetml" in resp.headers["content-type"]


def test_export_xlsx_content_disposition() -> None:
    """Content-Disposition 헤더에 lotto_draws.xlsx와 attachment가 포함된다."""
    with patch("lotto.web.data.get_draws", return_value=SAMPLE_DRAWS):
        resp = client.get("/api/export/xlsx")
    cd = resp.headers["content-disposition"]
    assert "lotto_draws.xlsx" in cd
    assert "attachment" in cd


def test_export_xlsx_readable() -> None:
    """응답 바이트가 유효한 xlsx 파일이다 (openpyxl로 열 수 있다)."""
    with patch("lotto.web.data.get_draws", return_value=SAMPLE_DRAWS):
        resp = client.get("/api/export/xlsx")
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    assert wb is not None


def test_export_xlsx_sheet_name() -> None:
    """시트 이름이 '로또당첨번호'이다."""
    with patch("lotto.web.data.get_draws", return_value=SAMPLE_DRAWS):
        resp = client.get("/api/export/xlsx")
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    assert "로또당첨번호" in wb.sheetnames


def test_export_xlsx_header_row() -> None:
    """첫 번째 행이 한국어 헤더 컬럼을 포함한다."""
    with patch("lotto.web.data.get_draws", return_value=SAMPLE_DRAWS):
        resp = client.get("/api/export/xlsx")
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    headers = [ws.cell(row=1, column=i).value for i in range(1, 12)]
    assert "회차" in headers
    assert "추첨일" in headers
    assert "보너스" in headers


def test_export_xlsx_row_count() -> None:
    """데이터 행 수 = 회차 수 (헤더 제외)."""
    with patch("lotto.web.data.get_draws", return_value=SAMPLE_DRAWS):
        resp = client.get("/api/export/xlsx")
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    assert ws.max_row - 1 == len(SAMPLE_DRAWS)


def test_export_xlsx_freeze_panes() -> None:
    """헤더 행이 고정되어 있다 (freeze_panes = A2)."""
    with patch("lotto.web.data.get_draws", return_value=SAMPLE_DRAWS):
        resp = client.get("/api/export/xlsx")
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    assert ws.freeze_panes == "A2"


def test_export_xlsx_empty_draws() -> None:
    """회차가 없으면 헤더 행만 포함된 xlsx가 반환된다."""
    with patch("lotto.web.data.get_draws", return_value=EMPTY_DRAWS):
        resp = client.get("/api/export/xlsx")
    assert resp.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    assert ws.max_row == 1
